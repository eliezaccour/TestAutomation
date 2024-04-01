import re

import pandas as pd
import json
import io
from datetime import datetime
import pytz
from openpyxl.styles import Font
from openpyxl.styles import Alignment
from openpyxl.styles import PatternFill
from openpyxl.cell.cell import ILLEGAL_CHARACTERS_RE

def initialize_st_tc_status(filename, sheetname, index_row, test_attendants, timezone, index_col_attendants=8, index_col_date=9, index_col_status=10):
    '''
    Initializes the status of the test case to "FAIL" and sets a timestamp in the Excel file.
    If the test case succeeds, the status field would be updated to PASS by the write_result_to_excel function.
    '''
    # Get current date and time
    curr_datetime = datetime.now(pytz.timezone(timezone))
    curr_datetime_ = curr_datetime.strftime("%d/%m/%Y %H:%M:%S")  # dd/mm/YY H:M:S
    # Write to file
    filename_ = filename
    options = {'strings_to_formulas': False, 'strings_to_urls': False}
    writer = pd.ExcelWriter(filename_, mode='a', engine='openpyxl', engine_kwargs={'options': options})
    worksheet = writer.sheets[sheetname]
    worksheet.cell(row=int(index_row), column=int(index_col_status)).value = "FAIL"
    worksheet.cell(row=int(index_row), column=int(index_col_date)).value = curr_datetime_
    worksheet.cell(row=int(index_row), column=int(index_col_attendants)).value = test_attendants
    worksheet.cell(row=int(index_row), column=int(index_col_status)).font = Font(name='Arial', size=10)
    worksheet.cell(row=int(index_row), column=int(index_col_date)).font = Font(name='Arial', size=10)
    worksheet.cell(row=int(index_row), column=int(index_col_attendants)).font = Font(name='Arial', size=10)
    worksheet.cell(row=int(index_row), column=int(index_col_status)).alignment = Alignment(horizontal='center', vertical='center')
    worksheet.cell(row=int(index_row), column=int(index_col_date)).alignment = Alignment(horizontal='center', vertical='center')
    worksheet.cell(row=int(index_row), column=int(index_col_attendants)).alignment = Alignment(horizontal='center', vertical='center')
    worksheet.cell(row=int(index_row), column=int(index_col_status)).fill = PatternFill(start_color="F4B084", end_color="F4B084", fill_type="solid")
    writer.save()

def write_st_result_to_excel(output, status, filename, sheetname, index_row, index_col_output=7, index_col_status=10):
    # Write to file
    filename_ = filename
    options = {'strings_to_formulas': False, 'strings_to_urls': False}
    writer = pd.ExcelWriter(filename_, mode='a', engine='openpyxl', engine_kwargs={'options': options})
    worksheet = writer.sheets[sheetname]
    worksheet.cell(row=int(index_row), column=int(index_col_output)).value = ILLEGAL_CHARACTERS_RE.sub(r'',output[:output.rfind('\n')])
    worksheet.cell(row=int(index_row), column=int(index_col_status)).value = status
    worksheet.cell(row=int(index_row), column=int(index_col_output)).font = Font(name='Courier New', size=9)
    worksheet.cell(row=int(index_row), column=int(index_col_status)).font = Font(name='Arial', size=10)
    worksheet.cell(row=int(index_row), column=int(index_col_output)).alignment = Alignment(horizontal='left', vertical='top', wrap_text=True)
    worksheet.cell(row=int(index_row), column=int(index_col_status)).alignment = Alignment(horizontal='center', vertical='center')
    worksheet.cell(row=int(index_row), column=int(index_col_status)).fill = PatternFill(start_color="A9D08E", end_color="A9D08E", fill_type = "solid")
    writer.save()

def read_config_key_value_pairs(filename, sheetname):
    config_keys = pd.read_excel(filename, sheet_name=sheetname)['Key']
    config_values = pd.read_excel(filename, sheet_name=sheetname)['Value']
    return config_keys, config_values

def read_input_command(filename, index, sheetname="1. System Test"):
    tc_row = pd.read_excel(filename, sheet_name=sheetname).loc[index-1]
    tc_command = tc_row['*** ${User_Input_Command} ***']
    tc_json = tc_row['*** ${User_Input_Parameters_JSON} ***']
    return tc_command, tc_json

def suppress_added_characters(output):
    '''
    The returned output has some unwanted added characters/stings ('[01;31m', '[K', and '[m'). This shall remove them.
    :param output: command output
    :return: fixed output format (as seen from command-line)
    '''
    output_ = output.replace('[01;31m', '')
    output__ = output_.replace('[K', '')
    output___ = output__.replace('[m', '')
    return output___


def populate_command_variables(command, user_input_json):
    """
    :param command: Command to populate [str].
    :param user_input_json: User input as a JSON string.
    :return: Populated command.
    """
    if user_input_json.startswith('{') and command != '_NA_':
        input_data = json.loads(user_input_json)
        for i in input_data:
            command = re.sub(r"\${"+i+"}", input_data[i], command)
    return command
